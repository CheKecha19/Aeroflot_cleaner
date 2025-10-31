# utils.py
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from pathlib import Path
from config import SHTAT_DIR, KONTUR_DIR, DIADOC_DIR, ONEC_DIR, MAX_FILE_AGE_DAYS, MAX_ROWS
from datetime import datetime, timedelta
import logging
logger = logging.getLogger(__name__)

def is_file_recent(file_path):
    """Проверяет, актуален ли файл (создан/изменен не более MAX_FILE_AGE_DAYS дней назад)"""
    if not file_path.exists():
        return False
    
    file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
    return (datetime.now() - file_mtime) <= timedelta(days=MAX_FILE_AGE_DAYS)

def find_latest_file(directory, patterns=None):
    """
    Находит самый новый файл в директории, соответствующий шаблонам
    patterns: список шаблонов, например ['*.xlsx', '*.xls']
    """
    if patterns is None:
        patterns = ['*.xlsx', '*.xls']
    
    files = []
    for pattern in patterns:
        try:
            pattern_files = list(directory.glob(pattern))
            for file in pattern_files:
                if is_file_recent(file):
                    files.append(file)
        except Exception as e:
            logger.debug(f"Ошибка при поиске файлов по шаблону {pattern} в {directory}: {e}")
    
    if not files:
        logger.debug(f"Не найдено актуальных файлов в {directory} по шаблонам {patterns}")
        return None
    
    latest_file = max(files, key=os.path.getmtime)
    logger.info(f"Выбран файл: {latest_file.name} (из {len(files)} найденных в {directory.name})")
    return latest_file

def get_onec_file():
    """Находит файл 1С"""
    return find_latest_file(ONEC_DIR)

def get_kontur_file():
    """Находит файл Контура"""
    return find_latest_file(KONTUR_DIR)

def get_diadoc_file():
    """Находит файл Диадока"""
    return find_latest_file(DIADOC_DIR)

def get_shtat_file():
    """Находит файл штатного расписания"""
    return find_latest_file(SHTAT_DIR)

def replace_yo(text):
    """Замена ё на е"""
    if pd.isna(text):
        return text
    return str(text).replace('ё', 'е').replace('Ё', 'Е')

def normalize_name(full_name):
    """Нормализация ФИО (извлечение имени и фамилии без отчества)"""
    if pd.isna(full_name):
        return ""
    
    name = replace_yo(str(full_name))
    parts = re.split(r'\s+', name.strip())
    
    if len(parts) >= 2:
        return f"{parts[0]} {parts[1]}".upper()
    elif len(parts) == 1:
        return parts[0].upper()
    return ""

def highlight_duplicates(df, column, duplicate_names, color='red'):
    """Подсветка дубликатов в DataFrame"""
    if color == 'red':
        fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    else:
        fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    for idx, row in df.iterrows():
        normalized_name = normalize_name(row[column])
        if normalized_name in duplicate_names:
            yield fill
        else:
            yield None

def save_with_formatting(df, filename, sheet_name, highlighting):
    """Сохранение DataFrame с форматированием"""
    df.to_excel(filename, sheet_name=sheet_name, index=False)
    
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    
    for idx, fill in enumerate(highlighting, 1):
        if fill:
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=idx).fill = fill
    
    wb.save(filename)

def load_kontur_data():
    """Загрузка данных из Контур Диадок"""
    try:
        kontur_file = get_kontur_file()
        if not kontur_file:
            logger.warning("Актуальный файл Контур Диадок не найден")
            return pd.DataFrame(columns=['Контур_Диадок_ФИО', 'Контур_Диадок_Администратор', 'Контур_Диадок_статус'])
        
        logger.info(f"Загрузка данных из файла: {kontur_file.name}")
        df = pd.read_excel(kontur_file)
        
        # Автоматически определяем структуру файла
        df = df.rename(columns={
            'ФИО': 'Контур_Диадок_ФИО',
            'Администратор': 'Контур_Диадок_Администратор',
            'Дата блокировки': 'Контур_Диадок_статус'
        })
        
        # Если переименование не сработало, ищем столбцы по содержимому
        if 'Контур_Диадок_ФИО' not in df.columns:
            for col in df.columns:
                if any(keyword in str(col).lower() for keyword in ['фио', 'ф.и.о.', 'name']):
                    df = df.rename(columns={col: 'Контур_Диадок_ФИО'})
                elif any(keyword in str(col).lower() for keyword in ['администратор', 'admin']):
                    df = df.rename(columns={col: 'Контур_Диадок_Администратор'})
                elif any(keyword in str(col).lower() for keyword in ['дата блокировки', 'блокировка', 'статус']):
                    df = df.rename(columns={col: 'Контур_Диадок_статус'})
        
        result_df = df[['Контур_Диадок_ФИО', 'Контур_Диадок_Администратор', 'Контур_Диадок_статус']].copy()
        
        # Преобразуем булевы значения
        if 'Контур_Диадок_Администратор' in result_df.columns:
            admin_series = result_df['Контур_Диадок_Администратор'].astype(str)
            admin_series = admin_series.apply(
                lambda x: 'да' if x.lower() in ['true', 'истина', '1', 'yes', 'да'] 
                else 'нет' if x.lower() in ['false', 'ложь', '0', 'no', 'нет'] 
                else x
            )
            result_df = result_df.assign(Контур_Диадок_Администратор=admin_series)
        
        # Преобразуем даты блокировки в статусы
        if 'Контур_Диадок_статус' in result_df.columns:
            status_series = result_df['Контур_Диадок_статус'].apply(
                lambda x: 'заблокирована' if pd.notna(x) and str(x).strip() != '' 
                else 'активна'
            )
            result_df = result_df.assign(Контур_Диадок_статус=status_series)
        
        logger.info(f"Загружено {len(result_df)} записей из Контур Диадок")
        return result_df
        
    except Exception as e:
        logger.error(f"Ошибка при загрузке данных Контур Диадок: {e}")
        return pd.DataFrame(columns=['Контур_Диадок_ФИО', 'Контур_Диадок_Администратор', 'Контур_Диадок_статус'])
    
def load_diadoc_data():
    """Загрузка данных из Сфера Курьер"""
    try:
        diadoc_file = get_diadoc_file()
        if not diadoc_file:
            logger.warning("Актуальный файл Сфера Курьер не найден")
            return pd.DataFrame(columns=['Сфера_Курьер_ФИО', 'Сфера_Курьер_Активен', 'Сфера_Курьер_Администратор'])
        
        logger.info(f"Загрузка данных из файла: {diadoc_file.name}")
        df = pd.read_excel(diadoc_file)
        
        # Автоматически определяем структуру файла
        df = df.rename(columns={
            'ФИО': 'Сфера_Курьер_ФИО',
            'Активен': 'Сфера_Курьер_Активен',
            'Администратор': 'Сфера_Курьер_Администратор'
        })
        
        # Если переименование не сработало, ищем столбцы по содержимому
        if 'Сфера_Курьер_ФИО' not in df.columns:
            for col in df.columns:
                if any(keyword in str(col).lower() for keyword in ['фио', 'ф.и.о.', 'name']):
                    df = df.rename(columns={col: 'Сфера_Курьер_ФИО'})
                elif any(keyword in str(col).lower() for keyword in ['активен', 'active', 'статус']):
                    df = df.rename(columns={col: 'Сфера_Курьер_Активен'})
                elif any(keyword in str(col).lower() for keyword in ['администратор', 'admin']):
                    df = df.rename(columns={col: 'Сфера_Курьер_Администратор'})
        
        result_df = df[['Сфера_Курьер_ФИО', 'Сфера_Курьер_Активен', 'Сфера_Курьер_Администратор']]
        logger.info(f"Загружено {len(result_df)} записей из Сфера Курьер")
        return result_df
    except Exception as e:
        logger.error(f"Ошибка при загрузке данных Сфера Курьер: {e}")
        return pd.DataFrame(columns=['Сфера_Курьер_ФИО', 'Сфера_Курьер_Активен', 'Сфера_Курьер_Администратор'])

def load_shtat_data():
    """Загрузка данных из штатного расписания"""
    try:
        shtat_file = get_shtat_file()
        if not shtat_file:
            logger.warning("Актуальный файл штатного расписания не найден")
            return pd.DataFrame(columns=['Штатное_ФИО'])
        
        logger.info(f"Загрузка данных из файла: {shtat_file.name}")
        df = pd.read_excel(shtat_file)
        
        # Автоматически определяем столбец с ФИО
        if 'Ф.И.О.' in df.columns:
            df = df.rename(columns={'Ф.И.О.': 'Штатное_ФИО'})
        elif 'ФИО' in df.columns:
            df = df.rename(columns={'ФИО': 'Штатное_ФИО'})
        else:
            # Ищем столбец с ФИО по содержимому
            for col in df.columns:
                if any(keyword in str(col).lower() for keyword in ['фио', 'ф.и.о.', 'фио сотрудника']):
                    df = df.rename(columns={col: 'Штатное_ФИО'})
                    break
        
        logger.info(f"Загружено {len(df)} записей из штатного расписания")
        return df[['Штатное_ФИО']]
    except Exception as e:
        logger.error(f"Ошибка при загрузке данных штатного расписания: {e}")
        return pd.DataFrame(columns=['Штатное_ФИО'])

def create_comparison_sheet(ad_employees, shtat_employees, filename):
    """Создание листа сравнения AD и Штатного расписания"""
    if not shtat_employees:
        return 0
    
    ad_set = set(normalize_name(name) for name in ad_employees)
    shtat_set = set(normalize_name(name) for name in shtat_employees)
    
    missing_in_shtat = ad_set - shtat_set
    
    comparison_data = []
    for name in missing_in_shtat:
        original_name = next((n for n in ad_employees if normalize_name(n) == name), name)
        comparison_data.append({
            'ФИО_AD': original_name,
            'Статус': 'Активен в AD, но отсутствует в штатном расписании'
        })
    
    comparison_df = pd.DataFrame(comparison_data)
    
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
        comparison_df.to_excel(writer, sheet_name='сравнение AD и Штатки', index=False)
    
    return len(missing_in_shtat)

def find_duplicates(df1, df2, col1, col2):
    """Поиск дубликатов между двумя DataFrame"""
    names1 = set(df1[col1].apply(normalize_name).dropna())
    names2 = set(df2[col2].apply(normalize_name).dropna())
    
    return names1.intersection(names2)

def find_internal_duplicates(df, column):
    """Поиск дубликатов внутри одного столбца"""
    normalized_names = df[column].apply(normalize_name)
    value_counts = normalized_names.value_counts()
    return set(value_counts[value_counts > 1].index)

def find_users_to_remove(edo_df, staff_df, gph_df):
    """Поиск пользователей для удаления из ЭДО"""
    all_valid_names = set()
    
    if not staff_df.empty and 'AD_ФИО' in staff_df.columns:
        all_valid_names.update(staff_df['AD_ФИО'].apply(normalize_name).dropna())
    
    if not gph_df.empty and 'AD_ФИО' in gph_df.columns:
        all_valid_names.update(gph_df['AD_ФИО'].apply(normalize_name).dropna())
    
    users_to_remove = []
    
    for _, row in edo_df.iterrows():
        fio_column = edo_df.columns[0]
        if pd.isna(row[fio_column]):
            continue
            
        normalized_name = normalize_name(row[fio_column])
        
        if normalized_name not in all_valid_names:
            if 'Контур_статус' in edo_df.columns and row['Контур_статус'] == 'активна':
                users_to_remove.append(row)
            elif 'Диадок_Активен' in edo_df.columns and row['Диадок_Активен'] == 'Да':
                users_to_remove.append(row)
            elif '1C_Активен' in edo_df.columns and row['1C_Активен'] == 'Да':
                users_to_remove.append(row)
    
    return pd.DataFrame(users_to_remove)